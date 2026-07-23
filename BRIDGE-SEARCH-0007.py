# BRIDGE-SEARCH-0007
from __future__ import annotations

import json
import math
import random
import re
import urllib.request
from pathlib import Path

from google.colab import userdata

BRIDGE_VERSION = "BRIDGE-SEARCH-0007"
BATCH_SIZE = 20
OWNER = "gear66me-ui"
REPOSITORY = "Galaxy_Viewer"
BRANCH = "sandbox"
REMOTE_FOLDER = "bridge-search"
TOKEN_SECRET = "GITHUB_TOKEN"
OUTPUT_PATH = Path(f"/content/{BRIDGE_VERSION}_{BATCH_SIZE}_GALAXIES.xlsx")
BASE_URL = "https://raw.githubusercontent.com/gear66me-ui/Galaxy_Viewer/79c24135206e87614000f9a67dc6447794a6a949/BRIDGE-SEARCH-0006.py"

FIELDS = [
    "Object",
    "Common name / nickname",
    "Alternate names",
    "Constellation",
    "ICRS coordinates",
    "Galaxy age",
    "Redshift (z) / Distance",
    "Morphological type",
    "Angular size",
    "Radial velocity",
    "Physical size",
    "Magnitude",
    "Data source",
]

with urllib.request.urlopen(BASE_URL, timeout=60) as response:
    base_source = response.read().decode("utf-8")

# Load the proven Bridge 6 search functions without launching its main routine.
base_namespace = {"__name__": "bridge_search_0007_base"}
exec(compile(base_source, "BRIDGE-SEARCH-0006-base.py", "exec"), base_namespace)
base_namespace["BATCH_SIZE"] = BATCH_SIZE
base_namespace["OUTPUT_PATH"] = OUTPUT_PATH

load_main = base_namespace["load_main"]
simbad_backup = base_namespace["simbad_backup"]
vizier_backup = base_namespace["vizier_backup"]
missing = base_namespace["missing"]
derive_size = base_namespace["derive_size"]
upload = base_namespace["upload"]


def parse_coords(text: str):
    try:
        parts = re.split(r"[\s,]+", str(text).strip())
        ra, dec = float(parts[0]), float(parts[1])
        if math.isfinite(ra) and math.isfinite(dec):
            return ra, dec
    except Exception:
        pass
    return None


def constellation_name(coord_text: str) -> str:
    coords = parse_coords(coord_text)
    if not coords:
        return "Constellation assignment pending coordinate verification"
    try:
        from astropy.coordinates import SkyCoord, get_constellation
        import astropy.units as u
        sky = SkyCoord(coords[0] * u.deg, coords[1] * u.deg, frame="icrs")
        return str(get_constellation(sky, short_name=False))
    except Exception:
        return "Constellation can be derived from the listed sky coordinates"


def clean_name(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def distinct_names(*values: str) -> list[str]:
    names = []
    seen = set()
    for value in values:
        text = clean_name(value)
        if not text:
            continue
        for token in re.split(r"\s{2,}|\s*[;,|]\s*", text):
            token = clean_name(token)
            key = token.lower()
            if token and key not in seen:
                seen.add(key)
                names.append(token)
    return names


def friendly_missing(field: str, constellation: str) -> str:
    messages = {
        "Common name / nickname": f"Catalog galaxy in {constellation} [descriptive label; no established nickname confirmed]",
        "Alternate names": "No additional catalog alias was confirmed in this search pass",
        "Galaxy age": "EST [observed galaxy; stellar-population age has not been measured directly and likely spans several billion years]",
        "Redshift (z) / Distance": "Published redshift or distance was not confirmed in this pass; deeper catalog measurements may exist",
        "Morphological type": "Galaxy morphology is not firmly classified; imaging suggests a cataloged extragalactic system",
        "Angular size": "Observed galaxy; catalog angular-diameter estimates vary by survey and measurement band",
        "Radial velocity": "Spectroscopic velocity was not confirmed in this pass; published measurements may vary by reference frame",
        "Physical size": "EST [physical extent awaits a securely adopted distance and angular diameter]",
        "Magnitude": "Observed galaxy; brightness estimates vary with filter, aperture, and catalog",
        "Data source": "Multiple astronomy catalogs and web-audit query; verify critical values",
    }
    return messages.get(field, "Observed galaxy; detailed catalog value remains to be verified")


def morphology_plain(value: str) -> str:
    text = clean_name(value)
    if not text or missing(text):
        return ""
    replacements = {
        "Galaxy [SIMBAD G]": "Galaxy; detailed morphology not specified [SIMBAD G]",
        "Galaxy classification [G]": "Galaxy; detailed morphology not specified [SIMBAD G]",
    }
    return replacements.get(text, text)


def to_viewer_row(primary: dict, backup_a: dict | None = None, backup_b: dict | None = None) -> dict[str, str]:
    backup_a = backup_a or {}
    backup_b = backup_b or {}
    coord = clean_name(primary.get("ICRS coordinates") or backup_a.get("ICRS coordinates") or backup_b.get("ICRS coordinates"))
    constellation = constellation_name(coord)
    names = distinct_names(primary.get("Object", ""), backup_a.get("Object", ""), backup_b.get("Object", ""))
    object_name = names[0] if names else "Catalog galaxy"
    aliases = [name for name in names[1:] if name.lower() != object_name.lower()]

    preferred_common = next((name for name in names if re.match(r"^(NGC|IC|UGC|Messier|M\s*\d+)", name, re.I)), "")
    common = preferred_common if preferred_common and preferred_common.lower() != object_name.lower() else ""

    def choose(field: str) -> str:
        values = [primary.get(field, ""), backup_a.get(field, ""), backup_b.get(field, "")]
        usable = [clean_name(v) for v in values if v and not missing(v)]
        if not usable:
            return friendly_missing(field, constellation)
        normalized = [re.sub(r"\[[^\]]*\]", "", v).strip().lower() for v in usable]
        for i, value in enumerate(normalized):
            if normalized.count(value) >= 2:
                return usable[i]
        return usable[0] if len(usable) == 1 else f"{usable[0]} [catalog values vary; alternate: {usable[1]}]"

    row = {
        "Object": object_name,
        "Common name / nickname": common or friendly_missing("Common name / nickname", constellation),
        "Alternate names": "; ".join(aliases) if aliases else friendly_missing("Alternate names", constellation),
        "Constellation": constellation,
        "ICRS coordinates": coord or "Coordinates require catalog verification",
        "Galaxy age": choose("Galaxy age"),
        "Redshift (z) / Distance": choose("Redshift (z) / Distance"),
        "Morphological type": morphology_plain(choose("Morphological type")),
        "Angular size": choose("Angular size"),
        "Radial velocity": choose("Radial velocity"),
        "Physical size": choose("Physical size"),
        "Magnitude": choose("Magnitude"),
        "Data source": "Multiple astronomy catalogs; web-audit query supplied for manual verification",
    }
    if "awaits" in row["Physical size"].lower():
        old_shape = {
            "Angular size": row["Angular size"],
            "Redshift (z) / Distance": row["Redshift (z) / Distance"],
            "Physical size": "",
        }
        derived = derive_size(old_shape)
        if derived and not missing(derived):
            row["Physical size"] = derived
    return row


def needs_backup(row: dict) -> bool:
    core = ["Redshift (z) / Distance", "Morphological type", "Angular size", "Radial velocity"]
    return any(missing(row.get(field, "")) for field in core)


def blank_old_row(row: dict) -> dict:
    old_fields = base_namespace["FIELDS"]
    out = {field: "" for field in old_fields}
    out["Object"] = row.get("Object", "")
    out["ICRS coordinates"] = row.get("ICRS coordinates", "")
    return out


def google_query(row: dict[str, str]) -> str:
    obj = row.get("Object", "galaxy")
    aliases = row.get("Alternate names", "")
    return (
        f'"{obj}" galaxy name nickname aliases morphology redshift distance radial velocity '
        f'magnitude angular size constellation SIMBAD NED HyperLEDA {aliases}'
    ).strip()


def build_google_audit(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    audit = []
    for row in rows:
        item = dict(row)
        item["Google search string"] = google_query(row)
        item["Manual verification status"] = "Pending manual review"
        item["Manual source notes"] = "Paste concise findings from Google, SIMBAD, NED, HyperLEDA, or another identified source"
        audit.append(item)
    return audit


def make_workbook(primary, backup, google_audit, combined):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="0B4F6C")
    header_font = Font(bold=True, color="FFFFFF")

    sheets = {
        "Primary Astronomy": (FIELDS, primary),
        "Backup Astronomy": (FIELDS, backup),
        "Google-Web Audit": (FIELDS + ["Google search string", "Manual verification status", "Manual source notes"], google_audit),
        "Combined": (FIELDS, combined),
        "Viewer Preview": (FIELDS, combined),
    }

    for title, (columns, rows) in sheets.items():
        ws = wb.create_sheet(title=title)
        ws.append(columns)
        for row in rows:
            ws.append([row.get(column, "") for column in columns])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for index, column in enumerate(columns, 1):
            width = 38 if column in {"Galaxy age", "Redshift (z) / Distance", "Morphological type", "Physical size", "Google search string", "Manual source notes"} else 24
            ws.column_dimensions[get_column_letter(index)].width = width

    wb.save(OUTPUT_PATH)


def display_preview(rows):
    from IPython.display import HTML, display

    payload = json.dumps(rows, ensure_ascii=False).replace("</", "<\\/")
    fields_payload = json.dumps(FIELDS, ensure_ascii=False)
    root = "bridge_search_0007_viewer"
    html = f"""
    <div id="{root}" style="background:#000;color:#7FDBFF;border:1px solid #0b4f6c;border-radius:10px;padding:14px;font-family:Arial,sans-serif;max-width:1180px;margin:10px auto;">
      <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;">
        <button id="{root}_back">◀ Back</button>
        <button id="{root}_forward">Forward ▶</button>
        <button id="{root}_copy">Copy Google Search</button>
        <button id="{root}_google">Open Google Search</button>
      </div>
      <div id="{root}_count" style="font-weight:bold;margin-bottom:8px;"></div>
      <div id="{root}_card"></div>
      <div style="margin-top:12px;font-weight:bold;">Google search string</div>
      <textarea id="{root}_query" readonly style="width:100%;min-height:76px;background:#071018;color:#d9f6ff;border:1px solid #0b4f6c;border-radius:6px;padding:8px;"></textarea>
      <div id="{root}_status" style="margin-top:6px;color:#9BE7FF;"></div>
    </div>
    <script>
    (() => {{
      const rows = {payload};
      const fields = {fields_payload};
      let index = 0;
      const root = "{root}";
      const card = document.getElementById(root + "_card");
      const count = document.getElementById(root + "_count");
      const queryBox = document.getElementById(root + "_query");
      const status = document.getElementById(root + "_status");
      function query(row) {{
        const aliases = row["Alternate names"] || "";
        return `"${{row["Object"] || "galaxy"}}" galaxy name nickname aliases morphology redshift distance radial velocity magnitude angular size constellation SIMBAD NED HyperLEDA ${{aliases}}`;
      }}
      function render() {{
        const row = rows[index];
        count.textContent = `Galaxy ${{index + 1}} of ${{rows.length}} — ${{row["Object"]}}`;
        card.innerHTML = "";
        const table = document.createElement("table");
        table.style.cssText = "width:100%;border-collapse:collapse";
        fields.forEach(field => {{
          const tr = document.createElement("tr");
          const th = document.createElement("th");
          const td = document.createElement("td");
          th.textContent = field;
          td.textContent = row[field];
          th.style.cssText = "text-align:left;vertical-align:top;width:220px;padding:7px;border-bottom:1px solid #12384a;color:#7FDBFF";
          td.style.cssText = "padding:7px;border-bottom:1px solid #12384a;color:#e8fbff;white-space:pre-wrap";
          tr.appendChild(th); tr.appendChild(td); table.appendChild(tr);
        }});
        card.appendChild(table);
        queryBox.value = query(row);
        status.textContent = "";
      }}
      document.getElementById(root + "_back").onclick = () => {{ index = (index - 1 + rows.length) % rows.length; render(); }};
      document.getElementById(root + "_forward").onclick = () => {{ index = (index + 1) % rows.length; render(); }};
      document.getElementById(root + "_copy").onclick = async () => {{
        try {{ await navigator.clipboard.writeText(queryBox.value); status.textContent = "Google search copied."; }}
        catch (error) {{ queryBox.focus(); queryBox.select(); status.textContent = "Search selected for manual copy."; }}
      }};
      document.getElementById(root + "_google").onclick = () => window.open("https://www.google.com/search?q=" + encodeURIComponent(queryBox.value), "_blank");
      render();
    }})();
    </script>
    """
    display(HTML(html))


def main():
    source_rows = load_main()

    backup_old = []
    for row in source_rows:
        first = simbad_backup(row) if needs_backup(row) else blank_old_row(row)
        merged = dict(row)
        for key, value in first.items():
            if value and not missing(value):
                merged[key] = value
        second = vizier_backup(merged) if needs_backup(merged) else blank_old_row(row)
        combined_backup = dict(first)
        for key, value in second.items():
            if value and not missing(value):
                combined_backup[key] = value
        backup_old.append(combined_backup)

    primary = [to_viewer_row(row) for row in source_rows]
    backup = [to_viewer_row(row, b) for row, b in zip(source_rows, backup_old)]
    combined = [to_viewer_row(row, b) for row, b in zip(source_rows, backup_old)]
    google_audit = build_google_audit(combined)

    # Guarantee that the viewer-facing Combined sheet contains no empty cells.
    for row in combined:
        constellation = row.get("Constellation", "the listed constellation")
        for field in FIELDS:
            if not clean_name(row.get(field, "")):
                row[field] = friendly_missing(field, constellation)

    make_workbook(primary, backup, google_audit, combined)

    token = userdata.get(TOKEN_SECRET)
    if not token:
        raise RuntimeError(f"Colab Secret {TOKEN_SECRET!r} is required")
    url = upload(OUTPUT_PATH, token)
    display_preview(combined)
    print(url)


if __name__ == "__main__":
    main()
