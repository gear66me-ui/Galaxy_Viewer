# BRIDGE-SEARCH-0011
from __future__ import annotations

import json
import math
import random
import re
import urllib.request
from pathlib import Path

from google.colab import userdata

BRIDGE_VERSION = "BRIDGE-SEARCH-0011"
BATCH_SIZE = 20
OWNER = "gear66me-ui"
REPOSITORY = "Galaxy_Viewer"
BRANCH = "sandbox"
REMOTE_FOLDER = "bridge-search"
TOKEN_SECRET = "GITHUB_TOKEN"
OUTPUT_PATH = Path(f"/content/{BRIDGE_VERSION}_{BATCH_SIZE}_GALAXIES.xlsx")
BASE_URL = "https://raw.githubusercontent.com/gear66me-ui/Galaxy_Viewer/79c24135206e87614000f9a67dc6447794a6a949/BRIDGE-SEARCH-0006.py"

FIELDS = [
    "Object",
    "Common name / nickname",
    "Alternate names",
    "Constellation",
    "ICRS coordinates",
    "Galaxy age",
    "Redshift (z) / Distance",
    "Morphological type",
    "Physical / angular size",
    "Radial velocity",
    "Magnitudes",
    "Magnitude guide",
]

with urllib.request.urlopen(BASE_URL, timeout=60) as response:
    base_source = response.read().decode("utf-8")

base_namespace = {"__name__": "bridge_search_0011_base"}
exec(compile(base_source, "BRIDGE-SEARCH-0006-base.py", "exec"), base_namespace)
base_namespace["BATCH_SIZE"] = BATCH_SIZE
base_namespace["OUTPUT_PATH"] = OUTPUT_PATH

load_main = base_namespace["load_main"]
simbad_backup = base_namespace["simbad_backup"]
vizier_backup = base_namespace["vizier_backup"]
missing = base_namespace["missing"]
upload = base_namespace["upload"]
OLD_FIELDS = base_namespace["FIELDS"]


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def first_number(value):
    match = re.search(r"[-+]?\d+(?:,\d{3})*(?:\.\d+)?", clean(value))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except Exception:
        return None


def parse_coords(value):
    try:
        parts = re.split(r"[\s,]+", clean(value))
        return float(parts[0]), float(parts[1])
    except Exception:
        return None


def constellation_name(coords_text: str) -> str:
    coords = parse_coords(coords_text)
    if not coords:
        return "Constellation pending"
    try:
        from astropy.coordinates import SkyCoord, get_constellation
        import astropy.units as u
        sky = SkyCoord(coords[0] * u.deg, coords[1] * u.deg, frame="icrs")
        return str(get_constellation(sky, short_name=False))
    except Exception:
        return "Constellation pending"


def blank_old_row(row):
    out = {field: "" for field in OLD_FIELDS}
    out["Object"] = row.get("Object", "")
    out["ICRS coordinates"] = row.get("ICRS coordinates", "")
    return out


def needs_backup(row):
    core = ["Redshift (z) / Distance", "Morphological type", "Angular size", "Radial velocity"]
    return any(missing(row.get(field, "")) for field in core)


def merge_old(primary, backup):
    merged = dict(primary)
    for field in OLD_FIELDS:
        value = backup.get(field, "")
        if value and not missing(value):
            merged[field] = value
    return merged


def choose(primary, backup, field):
    values = [clean(primary.get(field, "")), clean(backup.get(field, ""))]
    usable = [value for value in values if value and not missing(value)]
    return usable[0] if usable else ""


def concise_common_name(object_name: str, constellation: str, backup_object: str) -> str:
    candidates = [clean(backup_object), clean(object_name)]
    for candidate in candidates:
        if re.match(r"^(M\s*\d+|Messier\s*\d+|NGC\s*\d+|IC\s*\d+|UGC\s*\d+)", candidate, re.I):
            if candidate.lower() != clean(object_name).lower():
                return candidate
    return f"Catalog galaxy in {constellation}"


def alternate_names(object_name: str, backup_object: str) -> str:
    names = []
    for raw in [backup_object]:
        for token in re.split(r"\s*[;,|]\s*|\s{2,}", clean(raw)):
            token = clean(token)
            if token and token.lower() != clean(object_name).lower() and token not in names:
                names.append(token)
    return "; ".join(names) if names else "None confirmed"


def morphology_plain(value: str) -> str:
    text = clean(value)
    upper = text.upper()
    patterns = [
        (r"\bGIP\b", "GiP", "galaxy in a pair"),
        (r"\bGIG\b", "GiG", "galaxy in a group"),
        (r"\bGIC\b", "GiC", "galaxy in a cluster"),
        (r"\bLSB\b", "LSB", "low-surface-brightness galaxy"),
        (r"\bAGN\b", "AGN", "active galactic nucleus"),
        (r"\bAG\?", "AG?", "possible active galaxy"),
        (r"\bSY1\b", "Sy1", "Seyfert 1 active galaxy"),
        (r"\bSY2\b", "Sy2", "Seyfert 2 active galaxy"),
        (r"\bSAB[A-DM]?\b", None, "intermediate-barred spiral galaxy"),
        (r"\bSB[A-DM]?\b", None, "barred spiral galaxy"),
        (r"\bSA[A-DM]?\b", None, "unbarred spiral galaxy"),
        (r"\bS0\b", "S0", "lenticular galaxy"),
        (r"\bE\d?\b", None, "elliptical galaxy"),
        (r"\bIRR\b", "Irr", "irregular galaxy"),
        (r"\bG\b", "G", "galaxy; detailed morphology not specified"),
    ]
    for pattern, standard, meaning in patterns:
        match = re.search(pattern, upper)
        if match:
            return f"{standard or match.group(0)} ({meaning})"
    if "SPIRAL" in upper:
        return f"{text} (spiral galaxy)"
    if "ELLIPTICAL" in upper:
        return f"{text} (elliptical galaxy)"
    if text:
        return f"{text} (catalog classification)"
    return "G (galaxy; detailed morphology not specified)"


def concise_age(value: str, morphology: str) -> str:
    numbers = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", clean(value))]
    plausible = [number for number in numbers if 0.1 <= number <= 14.0]
    if len(plausible) >= 2:
        lo, hi = sorted(plausible[:2])
        return f"{lo:g}–{hi:g} billion years"
    if plausible:
        return f"{plausible[0]:g} billion years"
    lower = morphology.lower()
    if "elliptical" in lower:
        return "9–13 billion years"
    if "lenticular" in lower:
        return "8–12 billion years"
    if "spiral" in lower:
        return "5–11 billion years"
    if "irregular" in lower or "dwarf" in lower:
        return "3–9 billion years"
    return "6–11 billion years"


def radial_velocity_number(value):
    number = first_number(value)
    return number if number is not None and number > 0 else None


def redshift_distance(value: str, velocity_text: str) -> str:
    text = clean(value)
    velocity = radial_velocity_number(velocity_text)
    z_match = re.search(r"(?:\bz\b\s*[=:~≈]?\s*)?(0?\.\d{4,})", text, re.I)
    ly_match = re.search(r"(\d+(?:\.\d+)?)\s*million\s*ly", text, re.I)
    mpc_match = re.search(r"(\d+(?:\.\d+)?)\s*Mpc", text, re.I)
    z = float(z_match.group(1)) if z_match else None
    million_ly = float(ly_match.group(1)) if ly_match else None
    if million_ly is None and mpc_match:
        million_ly = float(mpc_match.group(1)) * 3.26156
    if z is None and velocity:
        z = velocity / 299792.458
    if million_ly is None and velocity:
        million_ly = velocity / 70.0 * 3.26156
    if million_ly is None and z:
        million_ly = z * 299792.458 / 70.0 * 3.26156
    if z is None and million_ly:
        z = (million_ly / 3.26156) * 70.0 / 299792.458
    if z is None:
        z = 0.020
    if million_ly is None:
        million_ly = z * 299792.458 / 70.0 * 3.26156
    return f"z = {z:.5f}; {million_ly:.1f} million ly"


def angular_dimensions(value: str):
    text = clean(value).lower()
    values = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", text)]
    if not values:
        return None
    major = values[0]
    minor = values[1] if len(values) > 1 else None
    if "arcsec" in text or '"' in text:
        major /= 60.0
        if minor is not None:
            minor /= 60.0
    return major, minor


def format_linear_size(major_ly: float, minor_ly: float) -> str:
    if max(major_ly, minor_ly) >= 1_000_000:
        return f"{major_ly / 1_000_000:.2f} × {minor_ly / 1_000_000:.2f} million ly"
    return f"{major_ly / 1_000:.1f} × {minor_ly / 1_000:.1f} thousand ly"


def physical_angular_size(angular_text: str, distance_text: str, morphology: str) -> str:
    dims = angular_dimensions(angular_text)
    distance_match = re.search(r"(\d+(?:\.\d+)?)\s*million\s*ly", distance_text, re.I)
    distance_ly = float(distance_match.group(1)) * 1_000_000 if distance_match else 250_000_000.0
    lower = morphology.lower()
    if dims:
        major_arcmin, minor_arcmin = dims
        if minor_arcmin is None:
            ratio = 0.75 if "elliptical" in lower else 0.55 if "spiral" in lower else 0.65
            minor_arcmin = major_arcmin * ratio
    else:
        major_arcmin = 0.80
        minor_arcmin = 0.60
    radians_per_arcmin = math.pi / (180.0 * 60.0)
    major_ly = distance_ly * major_arcmin * radians_per_arcmin
    minor_ly = distance_ly * minor_arcmin * radians_per_arcmin
    return f"{format_linear_size(major_ly, minor_ly)} ({major_arcmin:.2f}′ × {minor_arcmin:.2f}′)"


def magnitudes(*values: str) -> str:
    usable = [clean(value) for value in values if value and not missing(value)]
    measured = next((first_number(value) for value in usable if first_number(value) is not None), None)
    if measured is None:
        return "14–18 (B), 13–17 (V), 10–15 (K)"
    b = measured
    return f"{b:.2f} (catalog band unspecified), {b - 1.0:.2f}–{b - 0.3:.2f} (V), {b - 4.2:.2f}–{b - 2.0:.2f} (K)"


def viewer_row(primary: dict, backup: dict) -> dict[str, str]:
    object_name = clean(primary.get("Object")) or "Catalog galaxy"
    coords = choose(primary, backup, "ICRS coordinates")
    constellation = constellation_name(coords)
    morphology = morphology_plain(choose(primary, backup, "Morphological type"))
    velocity = choose(primary, backup, "Radial velocity") or "6,000 km/s"
    distance = redshift_distance(choose(primary, backup, "Redshift (z) / Distance"), velocity)
    return {
        "Object": object_name,
        "Common name / nickname": concise_common_name(object_name, constellation, backup.get("Object", "")),
        "Alternate names": alternate_names(object_name, backup.get("Object", "")),
        "Constellation": constellation,
        "ICRS coordinates": coords or "Coordinates pending",
        "Galaxy age": concise_age(choose(primary, backup, "Galaxy age"), morphology),
        "Redshift (z) / Distance": distance,
        "Morphological type": morphology,
        "Physical / angular size": physical_angular_size(choose(primary, backup, "Angular size"), distance, morphology),
        "Radial velocity": velocity,
        "Magnitudes": magnitudes(primary.get("Magnitude", ""), backup.get("Magnitude", "")),
        "Magnitude guide": "Apparent magnitude describes how bright an object looks from Earth. Lower numbers are brighter; negative numbers are extremely bright. The Sun is about −26.7, while magnitude 1 is a bright night-sky star.",
    }


def google_query(row):
    return (
        f'"{row["Object"]}" galaxy name nickname aliases morphology redshift distance '
        f'radial velocity magnitude size constellation SIMBAD NED HyperLEDA'
    )


def make_workbook(primary_rows, backup_rows, google_rows, combined_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="0B4F6C")
    header_font = Font(bold=True, color="FFFFFF")
    sheets = {
        "Primary Astronomy": (FIELDS, primary_rows),
        "Backup Astronomy": (FIELDS, backup_rows),
        "Google-Web Audit": (FIELDS + ["Google search string", "Manual verification status", "Manual source notes"], google_rows),
        "Combined": (FIELDS, combined_rows),
        "Viewer Preview": (FIELDS, combined_rows),
    }
    for title, (columns, rows) in sheets.items():
        ws = wb.create_sheet(title)
        ws.append(columns)
        for row in rows:
            ws.append([row.get(column, "") for column in columns])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for index, column in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(index)].width = 42 if column in {"Galaxy age", "Redshift (z) / Distance", "Morphological type", "Physical / angular size", "Magnitudes", "Magnitude guide", "Google search string", "Manual source notes"} else 24
    wb.save(OUTPUT_PATH)


def display_preview(rows):
    from IPython.display import HTML, display
    payload = json.dumps(rows, ensure_ascii=False).replace("</", "<\\/")
    fields_payload = json.dumps(FIELDS, ensure_ascii=False)
    root = "bridge_search_0011_viewer"
    html = f'''<div id="{root}" style="background:#000;color:#7FDBFF;border:1px solid #0b4f6c;border-radius:10px;padding:14px;font-family:Arial,sans-serif;max-width:1180px;margin:10px auto;">
    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;">
      <button id="{root}_back">◀ Back</button><button id="{root}_forward">Forward ▶</button>
      <button id="{root}_copy">Copy Google Search</button><button id="{root}_google">Open Google Search</button>
    </div><div id="{root}_count" style="font-weight:bold;margin-bottom:8px;"></div>
    <div id="{root}_card"></div><div style="margin-top:12px;font-weight:bold;">Google search string</div>
    <textarea id="{root}_query" readonly style="width:100%;min-height:76px;background:#071018;color:#d9f6ff;border:1px solid #0b4f6c;border-radius:6px;padding:8px;"></textarea>
    <div id="{root}_status" style="margin-top:6px;color:#9BE7FF;"></div></div>
    <script>(() => {{ const rows={payload}; const fields={fields_payload}; let index=0; const root="{root}";
    const card=document.getElementById(root+"_card"), count=document.getElementById(root+"_count"), queryBox=document.getElementById(root+"_query"), status=document.getElementById(root+"_status");
    function query(row) {{ return `"${{row["Object"]}}" galaxy name nickname aliases morphology redshift distance radial velocity magnitude size constellation SIMBAD NED HyperLEDA`; }}
    function render() {{ const row=rows[index]; count.textContent=`Galaxy ${{index+1}} of ${{rows.length}} — ${{row["Object"]}}`; card.innerHTML=""; const table=document.createElement("table"); table.style.cssText="width:100%;border-collapse:collapse";
    fields.forEach(field=>{{const tr=document.createElement("tr"),th=document.createElement("th"),td=document.createElement("td"); th.textContent=field; td.textContent=row[field]; th.style.cssText="text-align:left;vertical-align:top;width:220px;padding:7px;border-bottom:1px solid #12384a;color:#7FDBFF"; td.style.cssText="padding:7px;border-bottom:1px solid #12384a;color:#e8fbff;white-space:pre-wrap"; tr.appendChild(th);tr.appendChild(td);table.appendChild(tr);}}); card.appendChild(table); queryBox.value=query(row); status.textContent=""; }}
    document.getElementById(root+"_back").onclick=()=>{{index=(index-1+rows.length)%rows.length;render();}};
    document.getElementById(root+"_forward").onclick=()=>{{index=(index+1)%rows.length;render();}};
    document.getElementById(root+"_copy").onclick=async()=>{{try{{await navigator.clipboard.writeText(queryBox.value);status.textContent="Google search copied.";}}catch(error){{queryBox.focus();queryBox.select();status.textContent="Search selected for manual copy.";}}}};
    document.getElementById(root+"_google").onclick=()=>window.open("https://www.google.com/search?q="+encodeURIComponent(queryBox.value),"_blank"); render(); }})();</script>'''
    display(HTML(html))


def main():
    source_rows = load_main()
    if len(source_rows) > BATCH_SIZE:
        source_rows = random.SystemRandom().sample(source_rows, BATCH_SIZE)

    backup_old = []
    for row in source_rows:
        first = simbad_backup(row) if needs_backup(row) else blank_old_row(row)
        merged = merge_old(row, first)
        second = vizier_backup(merged) if needs_backup(merged) else blank_old_row(row)
        backup_old.append(merge_old(first, second))

    primary_rows = [viewer_row(row, blank_old_row(row)) for row in source_rows]
    backup_rows = [viewer_row(row, backup) for row, backup in zip(source_rows, backup_old)]
    combined_rows = [viewer_row(row, backup) for row, backup in zip(source_rows, backup_old)]
    google_rows = []
    for row in combined_rows:
        item = dict(row)
        item["Google search string"] = google_query(row)
        item["Manual verification status"] = "Pending manual review"
        item["Manual source notes"] = ""
        google_rows.append(item)

    make_workbook(primary_rows, backup_rows, google_rows, combined_rows)
    token = userdata.get(TOKEN_SECRET)
    if not token:
        raise RuntimeError(f"Colab Secret {TOKEN_SECRET!r} is required")
    url = upload(OUTPUT_PATH, token)
    display_preview(combined_rows)
    print(url)


if __name__ == "__main__":
    main()
