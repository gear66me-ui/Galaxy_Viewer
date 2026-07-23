# BRIDGE-SEARCH-0006
from __future__ import annotations

import re
import urllib.request

BRIDGE_VERSION = "BRIDGE-SEARCH-0006"
BASE_URL = "https://raw.githubusercontent.com/gear66me-ui/Galaxy_Viewer/c821f86a66b6d796a93a5cb04939e2f98b6e0a17/BRIDGE-SEARCH-0003.py"

with urllib.request.urlopen(BASE_URL, timeout=60) as response:
    source = response.read().decode("utf-8")

source = source.replace('BRIDGE_VERSION = "BRIDGE-SEARCH-0003"', 'BRIDGE_VERSION = "BRIDGE-SEARCH-0006"')
source = source.replace('BATCH_SIZE = 500', 'BATCH_SIZE = 10')
source = source.replace('def fetch_text(url: str, timeout: int = 120)', 'def fetch_text(url: str, timeout: int = 45)')
source = source.replace('timeout=90))', 'timeout=20))')
source = source.replace('import re\n', 'import re\nimport random\n')
source = source.replace(
    'return [{field: str(row.get(field, "") or "") for field in FIELDS} for row in rows[:BATCH_SIZE]]',
    'selected = random.SystemRandom().sample(rows, BATCH_SIZE)\n    return [{field: str(row.get(field, "") or "") for field in FIELDS} for row in selected]'
)

new_workbook = r'''def make_workbook(sheets: dict[str, list[dict[str, str]]]):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError("openpyxl is required to generate the workbook") from exc

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    header_fill = PatternFill("solid", fgColor="0B4F6C")
    header_font = Font(bold=True, color="FFFFFF")

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(FIELDS)
        for row in rows:
            ws.append([row.get(field, "") for field in FIELDS])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col_index, field in enumerate(FIELDS, start=1):
            width = 34 if field in {
                "Galaxy age", "Redshift (z) / Distance", "Morphological type",
                "Physical size", "Data source", "Data score"
            } else 22
            ws.column_dimensions[get_column_letter(col_index)].width = width

    wb.save(OUTPUT_PATH)
'''

source, count = re.subn(
    r'def make_workbook\(sheets: dict\[str, list\[dict\[str, str\]\]\]\):.*?\n\ndef upload\(',
    new_workbook + '\n\ndef upload(',
    source,
    flags=re.S,
)
if count != 1:
    raise RuntimeError("Could not replace the BRIDGE-SEARCH-0003 workbook exporter")

new_main = r'''def _blank_backup(row):
    out = {field: "" for field in FIELDS}
    out["Object"] = row.get("Object", "")
    out["ICRS coordinates"] = row.get("ICRS coordinates", "")
    return out


def _needs_backup(row):
    critical = [
        "Redshift (z) / Distance",
        "Morphological type",
        "Angular size",
        "Radial velocity",
    ]
    return any(missing(row.get(field, "")) for field in critical)


def _merge_nonempty(primary, backup):
    merged = dict(primary)
    for field in FIELDS:
        value = backup.get(field, "")
        if value and not missing(value):
            merged[field] = value
    return merged


def display_viewer_preview(rows):
    import json
    from IPython.display import HTML, display

    payload = json.dumps(rows, ensure_ascii=False).replace("</", "<\\/")
    fields_payload = json.dumps(FIELDS, ensure_ascii=False)
    root_id = "bridge_search_0006_viewer"

    html_block = f"""
    <div id="{root_id}" style="background:#000;color:#7FDBFF;border:1px solid #0b4f6c;border-radius:10px;padding:14px;font-family:Arial,sans-serif;max-width:1180px;margin:10px auto;">
      <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;">
        <button id="{root_id}_back">◀ Back</button>
        <button id="{root_id}_forward">Forward ▶</button>
        <button id="{root_id}_copy">Copy Google Search</button>
        <button id="{root_id}_google">Open Google Search</button>
      </div>
      <div id="{root_id}_count" style="font-weight:bold;margin-bottom:8px;"></div>
      <div id="{root_id}_card"></div>
      <div style="margin-top:12px;font-weight:bold;">Google search string</div>
      <textarea id="{root_id}_query" readonly style="width:100%;min-height:72px;background:#071018;color:#d9f6ff;border:1px solid #0b4f6c;border-radius:6px;padding:8px;"></textarea>
      <div id="{root_id}_status" style="margin-top:6px;color:#9BE7FF;"></div>
    </div>
    <script>
    (() => {{
      const rows = {payload};
      const fields = {fields_payload};
      let index = 0;
      const root = "{root_id}";
      const card = document.getElementById(root + "_card");
      const count = document.getElementById(root + "_count");
      const queryBox = document.getElementById(root + "_query");
      const status = document.getElementById(root + "_status");

      function googleQuery(row) {{
        const name = row["Object"] || "galaxy";
        return `"${{name}}" galaxy morphology redshift distance radial velocity magnitude angular size SIMBAD NED astronomy`;
      }}

      function render() {{
        const row = rows[index];
        count.textContent = `Galaxy ${{index + 1}} of ${{rows.length}} — ${{row["Object"] || "Unnamed"}}`;
        card.innerHTML = "";
        const table = document.createElement("table");
        table.style.width = "100%";
        table.style.borderCollapse = "collapse";
        fields.forEach(field => {{
          const tr = document.createElement("tr");
          const th = document.createElement("th");
          const td = document.createElement("td");
          th.textContent = field;
          td.textContent = row[field] || "Not available";
          th.style.cssText = "text-align:left;vertical-align:top;width:220px;padding:7px;border-bottom:1px solid #12384a;color:#7FDBFF;";
          td.style.cssText = "padding:7px;border-bottom:1px solid #12384a;color:#e8fbff;white-space:pre-wrap;";
          tr.appendChild(th);
          tr.appendChild(td);
          table.appendChild(tr);
        }});
        card.appendChild(table);
        queryBox.value = googleQuery(row);
        status.textContent = "";
      }}

      document.getElementById(root + "_back").onclick = () => {{
        index = (index - 1 + rows.length) % rows.length;
        render();
      }};
      document.getElementById(root + "_forward").onclick = () => {{
        index = (index + 1) % rows.length;
        render();
      }};
      document.getElementById(root + "_copy").onclick = async () => {{
        try {{
          await navigator.clipboard.writeText(queryBox.value);
          status.textContent = "Google search copied.";
        }} catch (error) {{
          queryBox.focus();
          queryBox.select();
          status.textContent = "Clipboard access was blocked. The search text is selected for manual copy.";
        }}
      }};
      document.getElementById(root + "_google").onclick = () => {{
        window.open("https://www.google.com/search?q=" + encodeURIComponent(queryBox.value), "_blank");
      }};
      render();
    }})();
    </script>
    """
    display(HTML(html_block))


def main():
    source_rows = load_main()

    backup1 = []
    for row in source_rows:
        backup1.append(simbad_backup(row) if _needs_backup(row) else _blank_backup(row))

    backup2 = []
    for row, first_backup in zip(source_rows, backup1):
        merged = _merge_nonempty(row, first_backup)
        backup2.append(vizier_backup(merged) if _needs_backup(merged) else _blank_backup(row))

    combined = build_combined(source_rows, backup1, backup2)
    make_workbook({
        "Main": source_rows,
        "Backup 1": backup1,
        "Backup 2": backup2,
        "Combined": combined,
        "Viewer Preview": combined,
    })

    token = userdata.get(TOKEN_SECRET)
    if not token:
        raise RuntimeError(f"Colab Secret {TOKEN_SECRET!r} is required")

    url = upload(OUTPUT_PATH, token)
    display_viewer_preview(combined)
    print(url)


if __name__ == "__main__":
    main()
'''

source, count = re.subn(
    r'def main\(\):.*?if __name__ == "__main__":\n    main\(\)',
    new_main,
    source,
    flags=re.S,
)
if count != 1:
    raise RuntimeError("Could not replace the BRIDGE-SEARCH-0003 main routine")

exec(compile(source, "BRIDGE-SEARCH-0006-runtime.py", "exec"), globals())
